import gi
from openpyxl import Workbook

gi.require_version('Gtk', '3.0')
from gi.repository import Gtk

from openpyxl.styles import Font, Alignment, Side, Border, PatternFill

# whatis = lambda obj: print(type(obj), "\n\t" + "\n\t".join(dir(obj)))

from openpyxl.styles.borders import BORDER_THIN


def create_a_report(lStore_now, lStore_medium, lStore_perspective, place):
    wb = Workbook()  # создаем рабочую книгу, в которую будем сохранять данные из workbook
    ws1 = wb.active  # создаем рабочий лист
    ws1.sheet_properties.tabColor = "1072BA"  # задаем цвет вкладки
    ws1.title = "Таблица срочных дел"  # задаем имя вкладки
    ws2 = wb.create_sheet()  # создаем рабочий лист
    ws2.sheet_properties.tabColor = "1072BA"  # задаем цвет вкладки
    ws2.title = "Таблица среднесрочных дел"  # задаем имя вкладки
    ws3 = wb.create_sheet()  # создаем рабочий лист
    ws3.sheet_properties.tabColor = "1072BA"  # задаем цвет вкладки
    ws3.title = "Таблица перспективных дел"  # задаем имя вкладки
    # задаем наименование столбцов
    ws = [ws1, ws2, ws3]
    for w in ws:
        w['A1'] = "№"
        w['B1'] = "Заметка"
        w['C1'] = "Дата"
        w.column_dimensions['A'].width = 6
        w.column_dimensions['B'].width = 60
        w.column_dimensions['C'].width = 20

    thin_border = Border(  # выделение границ ячеек
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

    for row in range(2, len(lStore_now) + 2):  # цикл по строкам с данными, начиная с 50-й недели
        path = Gtk.TreePath(row - 2)  # перебор строк с присвоением в path
        treeiter = lStore_now.get_iter(path)  # получение iter, соответствующее path
        ws1.cell(row, 1).value = lStore_now.get_value(treeiter, 0)  # присваивает ключу словаря значение столбца id
        ws1.cell(row, 2).value = lStore_now.get_value(treeiter, 1)
        ws1.cell(row, 3).value = lStore_now.get_value(treeiter, 2)

    for row in range(2, len(lStore_medium) + 2):  # цикл по строкам с данными, начиная с 50-й недели
        path = Gtk.TreePath(row - 2)  # перебор строк с присвоением в path
        treeiter = lStore_medium.get_iter(path)  # получение iter, соответствующее path
        ws2.cell(row, 1).value = lStore_medium.get_value(treeiter, 0)  # присваивает ключу словаря значение столбца id
        ws2.cell(row, 2).value = lStore_medium.get_value(treeiter, 1)
        ws2.cell(row, 3).value = lStore_medium.get_value(treeiter, 2)

    for row in range(2, len(lStore_perspective) + 2):  # цикл по строкам с данными, начиная с 50-й недели
        path = Gtk.TreePath(row - 2)  # перебор строк с присвоением в path
        treeiter = lStore_perspective.get_iter(path)  # получение iter, соответствующее path
        ws3.cell(row, 1).value = lStore_perspective.get_value(treeiter,
                                                              0)  # присваивает ключу словаря значение столбца id
        ws3.cell(row, 2).value = lStore_perspective.get_value(treeiter, 1)
        ws3.cell(row, 3).value = lStore_perspective.get_value(treeiter, 2)

    for w in ws:
        # цикл для задания ячейкам заголовков свойств
        for row in w.iter_cols(min_col=1, max_col=3, min_row=1, max_row=1):
            for cel in row:
                cel.font = Font(size=12, bold=True)  # размер шрифта и жирное выделение
                cel.alignment = Alignment(horizontal="center", vertical="center",
                                          wrapText=True)  # выравнивание по центру и разрешение переноса строк
                cel.fill = PatternFill(start_color="FFFAFA", end_color="FFFAFA", fill_type="solid")

        # цикл для выделения границ ячеек
        for row in w.iter_cols(min_col=1, max_col=3, min_row=1, max_row=w.max_row):
            for cel in row:
                cel.border = thin_border
    try:
        wb.save(place + '\\ToDoIt.xlsx')  # сохранение таблицы в указанную директорию
        return "OK"
    except:
        return "Error"
    # whatis(ws1.column_dimensions['A'].width)
